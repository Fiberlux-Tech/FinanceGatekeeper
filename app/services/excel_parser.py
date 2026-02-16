# app/services/excel_parser.py
# (This file is responsible for all Excel file ingestion and parsing.)

"""
Excel Parser Service.

Refactored from legacy Flask-coupled function to a dependency-injected service class.

Stripped:
    - from flask import current_app
    - @require_jwt decorator
    - Inner helper functions (now module-level with type hints)
    - print() statements (replaced with structured logging)

Injected:
    - VariableService (for master variable lookups)
    - logging.Logger (via BaseService)
    - AppConfig (via get_config())

Chain of Custody:
    - SHA-256 hash computed at file ingestion boundary (CLAUDE.md mandate).

Defensive File Handling:
    - PermissionError caught and surfaced as user-friendly ServiceResult.
"""

from __future__ import annotations

import hashlib
import traceback
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import AppConfig, get_config
from app.logger import StructuredLogger
from app.models.enums import FileStatus
from app.models.service_models import ServiceResult
from app.services.base_service import BaseService
from app.services.financial_engine import calculate_financial_metrics
from app.services.variables import VariableService
from app.utils.audit import log_audit_event
from app.utils.general import convert_to_json_safe
from app.utils.string_helpers import normalize_keys


# ---------------------------------------------------------------------------
# Module-level type-safe conversion helpers
# ---------------------------------------------------------------------------

def safe_decimal(val: Union[int, float, str, None], logger: StructuredLogger) -> Decimal:
    """Convert a value to Decimal, treating None, empty strings, and invalid values as Decimal("0").

    When using ``data_only=True``, openpyxl may return Excel error strings
    (#VALUE!, #DIV/0!, etc.) instead of computed values.  These are detected
    and logged so broken templates can be identified.

    Args:
        val: The raw cell value from openpyxl.
        logger: Logger instance for warning on conversion failures.

    Returns:
        The Decimal representation of *val*, or ``Decimal("0")`` on failure.
    """
    if val is not None and val != '':
        # Check for Excel error strings
        if isinstance(val, str) and val.startswith('#'):
            logger.warning("Excel error detected in cell: %s - Template may be broken", val)
            return Decimal("0")

        try:
            return Decimal(str(val))
        except (ValueError, TypeError, ArithmeticError):
            logger.warning(
                "Failed to convert value to Decimal: %s (type: %s)", val, type(val).__name__,
            )
            return Decimal("0")
    return Decimal("0")


def safe_int(val: Union[int, float, str, None]) -> int:
    """Convert a value to int, treating None, empty strings, and invalid values as 0.

    Handles string representations of floats (e.g. ``"5.0"`` -> ``5``) which
    are common in Excel exports.

    Args:
        val: The raw cell value from openpyxl.

    Returns:
        The integer representation of *val*, or ``0`` on failure.
    """
    if val is not None and val != '':
        try:
            return int(float(val))  # Handle "5.0" string -> 5
        except (ValueError, TypeError):
            return 0
    return 0


# ---------------------------------------------------------------------------
# Module-level field classification for header cell parsing
# Sourced from AppConfig (single source of truth — M6).
# Imported at module level to avoid repeated attribute lookups in hot loops.
# ---------------------------------------------------------------------------
_cfg: AppConfig = get_config()
_DECIMAL_FIELDS: frozenset[str] = _cfg.DECIMAL_FIELDS
_INT_FIELDS: frozenset[str] = _cfg.INT_FIELDS
_BOOL_FIELDS: frozenset[str] = _cfg.BOOL_FIELDS
_NUMERIC_FIELDS: frozenset[str] = _DECIMAL_FIELDS | _INT_FIELDS


def _compute_sha256(file_stream: BinaryIO) -> str:
    """Compute the SHA-256 hash of a file stream.

    The stream position is reset to the beginning after hashing so
    subsequent reads start from byte zero.

    Args:
        file_stream: A seekable binary file-like object.

    Returns:
        Hex-encoded SHA-256 digest string.
    """
    file_stream.seek(0)
    sha256_hash = hashlib.sha256()
    while True:
        chunk: bytes = file_stream.read(8192)
        if not chunk:
            break
        sha256_hash.update(chunk)
    file_stream.seek(0)
    return sha256_hash.hexdigest()


# ---------------------------------------------------------------------------
# Service class
# ---------------------------------------------------------------------------

class ExcelParserService(BaseService):
    """Service for ingesting and parsing Excel financial templates.

    Orchestrates reading, validating, and calculating data from the uploaded
    Excel file, using master variables for key financial rates.

    Dependencies are injected via ``__init__`` -- no Flask globals are used.
    """

    def __init__(
        self,
        variable_service: VariableService,
        config: AppConfig,
        logger: StructuredLogger,
        file_guards: Optional["FileGuardsService"] = None,
    ) -> None:
        """Initialise the parser with its runtime dependencies.

        Args:
            variable_service: Service used to retrieve master variable rates.
            config: Application configuration (injected, not singleton).
            logger: Logger instance for structured logging.
            file_guards: Optional file guards service for path-based safety
                checks.  When ``None``, path-based methods skip the
                readiness check and open the file directly.
        """
        super().__init__(logger)
        self._variable_service: VariableService = variable_service
        self._config: AppConfig = config
        # Lazy import type only — avoids circular import at module level.
        from app.services.file_guards import FileGuardsService as _FGS
        self._file_guards: Optional[_FGS] = file_guards

    # ------------------------------------------------------------------
    # Type aliases for internal data structures
    # ------------------------------------------------------------------
    _RowDict = dict[str, Union[int, float, Decimal, str, None]]
    _HeaderDict = dict[str, Union[Decimal, float, str, bool, None, dict[str, Union[Decimal, float, str, None]]]]
    _MasterRates = dict[str, Optional[Decimal]]
    _MasterSnapshot = dict[str, Union[Decimal, str, None]]

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def process_excel_file(self, excel_file: BinaryIO) -> ServiceResult:
        """Parse the uploaded Excel template and return calculated financial data.

        Orchestrates six decomposed steps:
            1. ``_hash_and_audit`` — SHA-256 chain of custody.
            2. ``_fetch_master_variables`` — rate validation & snapshot.
            3. ``_extract_header_data`` — header cell reads + master injection.
            4. ``_extract_table_rows`` — generic row-iteration (services & costs).
            5. ``_transform_and_enrich`` — currency tagging, preview totals, validation.
            6. ``_assemble_response`` — financial engine + JSON assembly.

        Args:
            excel_file: A seekable binary file-like object containing the
                Excel workbook (.xlsx).

        Returns:
            A ``ServiceResult`` with ``success=True`` and the parsed data
            package on success, or ``success=False`` with an error description
            on failure.
        """
        try:
            # Step 1: Chain of custody
            file_hash: str = self._hash_and_audit(excel_file)

            # Step 2: Master variable rates
            var_result: Union[ServiceResult, tuple[_MasterRates, _MasterSnapshot]] = (
                self._fetch_master_variables()
            )
            if isinstance(var_result, ServiceResult):
                return var_result
            latest_rates, snapshot = var_result

            # Step 3 & 4: Open workbook and extract all worksheet data
            self._logger.info("Reading Excel file with openpyxl (read_only mode for memory optimization)")
            excel_file.seek(0)

            workbook: Optional[Workbook] = None
            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)
                worksheet: Worksheet = workbook[self._config.PLANTILLA_SHEET_NAME]

                self._logger.info(
                    "Excel sheet loaded: %s rows x %s columns",
                    worksheet.max_row,
                    worksheet.max_column,
                )

                header_data: _HeaderDict = self._extract_header_data(
                    worksheet, latest_rates, snapshot,
                )

                recurring_services_data: list[_RowDict] = self._extract_table_rows(
                    worksheet,
                    self._config.RECURRING_SERVICES_START_ROW,
                    self._config.RECURRING_SERVICES_COLUMNS,
                )
                self._logger.info("SUCCESS: Read %d recurring service records", len(recurring_services_data))

                fixed_costs_data: list[_RowDict] = self._extract_table_rows(
                    worksheet,
                    self._config.FIXED_COSTS_START_ROW,
                    self._config.FIXED_COSTS_COLUMNS,
                )
                self._logger.debug("Read %d fixed cost records", len(fixed_costs_data))
            finally:
                if workbook:
                    workbook.close()
                    self._logger.info("Workbook closed successfully")

            # Step 5: Enrich and validate
            enrich_result: Union[ServiceResult, dict[str, object]] = (
                self._transform_and_enrich(header_data, recurring_services_data, fixed_costs_data)
            )
            if isinstance(enrich_result, ServiceResult):
                return enrich_result

            # Step 6: Financial metrics + response assembly
            return self._assemble_response(
                header_data, enrich_result, fixed_costs_data, recurring_services_data, file_hash,
            )

        except PermissionError as perm_err:
            self._logger.error(
                "PermissionError during Excel processing: %s\n%s",
                perm_err,
                traceback.format_exc(),
            )
            return ServiceResult(
                success=False,
                error=(
                    "The Excel file could not be accessed because it is locked by another "
                    "application. Please close any programs that may have the file open "
                    "and try again."
                ),
                status_code=423,
            )

        except Exception as exc:
            self._logger.error(
                "Unexpected error during Excel processing: %s\n%s",
                exc,
                traceback.format_exc(),
            )
            return ServiceResult(
                success=False,
                error=f"An unexpected error occurred: {str(exc)}",
                status_code=500,
            )

    def extract_metadata(self, file_path: Path) -> ServiceResult:
        """Extract header metadata from a local Excel file (lightweight).

        Reads only the header cells (C2–C10) defined in
        ``config.VARIABLES_TO_EXTRACT`` without parsing the full table
        rows or running the financial engine.  Designed for fast card
        population in the Card Engine (~50 ms per file).

        Runs ``FileGuardsService.check_file_status`` before opening
        the file when *file_guards* was provided at construction time.

        Args:
            file_path: Absolute path to a ``.xlsx`` file in the inbox.

        Returns:
            ``ServiceResult`` with ``data`` containing a ``dict`` of
            header field values on success, or an error on failure.
        """
        try:
            # Safety check (when file guards are available)
            if self._file_guards is not None:
                check = self._file_guards.check_file_status(file_path)
                if check.status != FileStatus.READY:
                    return ServiceResult(
                        success=False,
                        error=check.message,
                        status_code=423,
                    )

            workbook: Optional[Workbook] = None
            try:
                workbook = load_workbook(
                    file_path, read_only=True, data_only=True,
                )
                worksheet: Worksheet = workbook[self._config.PLANTILLA_SHEET_NAME]

                header_data: dict[str, Union[int, float, str]] = {}
                for var_name, cell_ref in self._config.VARIABLES_TO_EXTRACT.items():
                    cell_value = worksheet[cell_ref].value
                    if var_name in _DECIMAL_FIELDS:
                        header_data[var_name] = safe_decimal(cell_value, self._logger)
                    elif var_name in _INT_FIELDS:
                        header_data[var_name] = safe_int(cell_value)
                    elif var_name in _BOOL_FIELDS:
                        header_data[var_name] = bool(cell_value) if cell_value is not None else False
                    else:
                        header_data[var_name] = (
                            str(cell_value) if cell_value is not None else ""
                        )

                header_data = normalize_keys(header_data)
            finally:
                if workbook:
                    workbook.close()

            self._logger.debug(
                "Metadata extracted from %s: %s",
                file_path.name,
                {k: v for k, v in header_data.items() if k in ("client_name", "salesman", "unidad_negocio")},
            )
            return ServiceResult(success=True, data=header_data)

        except PermissionError:
            self._logger.warning(
                "File locked during metadata extraction: %s", file_path.name,
            )
            return ServiceResult(
                success=False,
                error=(
                    "The file is locked by another application. "
                    "Please close Excel and try again."
                ),
                status_code=423,
            )

        except KeyError:
            self._logger.warning(
                "Missing worksheet '%s' in %s",
                self._config.PLANTILLA_SHEET_NAME,
                file_path.name,
            )
            return ServiceResult(
                success=False,
                error=(
                    f"Invalid template: worksheet "
                    f"'{self._config.PLANTILLA_SHEET_NAME}' not found."
                ),
                status_code=400,
            )

        except Exception as exc:
            self._logger.error(
                "Metadata extraction failed for %s: %s\n%s",
                file_path.name,
                exc,
                traceback.format_exc(),
            )
            return ServiceResult(
                success=False,
                error=f"Failed to read file: {exc}",
                status_code=500,
            )

    def process_local_file(self, file_path: Path) -> ServiceResult:
        """Parse a local Excel file through the full financial pipeline.

        Equivalent to ``process_excel_file()`` but accepts a filesystem
        ``Path`` instead of a ``BinaryIO`` stream.  Opens the file,
        delegates to the existing 6-step pipeline, and returns the
        same ``ServiceResult`` data package.

        Args:
            file_path: Absolute path to a ``.xlsx`` file.

        Returns:
            ``ServiceResult`` with the full financial data package on
            success, or an error on failure.
        """
        try:
            # Safety check
            if self._file_guards is not None:
                check = self._file_guards.check_file_status(file_path)
                if check.status != FileStatus.READY:
                    return ServiceResult(
                        success=False,
                        error=check.message,
                        status_code=423,
                    )

            with open(file_path, "rb") as fh:
                return self.process_excel_file(fh)

        except PermissionError:
            self._logger.warning(
                "File locked during full parse: %s", file_path.name,
            )
            return ServiceResult(
                success=False,
                error=(
                    "The file is locked by another application. "
                    "Please close Excel and try again."
                ),
                status_code=423,
            )

        except Exception as exc:
            self._logger.error(
                "process_local_file failed for %s: %s\n%s",
                file_path.name,
                exc,
                traceback.format_exc(),
            )
            return ServiceResult(
                success=False,
                error=f"An unexpected error occurred: {exc}",
                status_code=500,
            )

    # ------------------------------------------------------------------
    # Private helpers — each encapsulates one responsibility
    # ------------------------------------------------------------------

    def _hash_and_audit(self, excel_file: BinaryIO) -> str:
        """Compute SHA-256 hash at the ingestion boundary and log audit event.

        Args:
            excel_file: A seekable binary file-like object.

        Returns:
            Hex-encoded SHA-256 digest string.
        """
        file_hash: str = _compute_sha256(excel_file)
        self._logger.info("File ingested. SHA-256: %s", file_hash)

        log_audit_event(
            logger=self._logger,
            action="INGEST_EXCEL",
            entity_type="ExcelFile",
            entity_id=file_hash,
            user_id="system",
            details={"sha256": file_hash, "timestamp": datetime.now(timezone.utc).isoformat()},
        )
        return file_hash

    def _fetch_master_variables(
        self,
    ) -> Union[ServiceResult, tuple[_MasterRates, _MasterSnapshot]]:
        """Fetch, validate, and snapshot master variable rates.

        Returns:
            On success: a tuple of (normalized_rates, frozen_snapshot).
            On failure: a ``ServiceResult`` with an error message.
        """
        required_keys: list[str] = ['tipoCambio', 'costoCapital', 'tasaCartaFianza']
        latest_rates: _MasterRates = (
            self._variable_service.get_latest_master_variables(required_keys)
        )
        latest_rates = normalize_keys(latest_rates)

        if (
            latest_rates.get('tipo_cambio') is None
            or latest_rates.get('costo_capital') is None
            or latest_rates.get('tasa_carta_fianza') is None
        ):
            return ServiceResult(
                success=False,
                error=(
                    "Cannot calculate financial metrics. System rates (Tipo de Cambio, "
                    "Costo Capital, or Tasa Carta Fianza) are missing. Please ensure they "
                    "have been set by the Finance department."
                ),
                status_code=400,
            )

        snapshot: _MasterSnapshot = {
            'tipo_cambio': latest_rates['tipo_cambio'],
            'costo_capital': latest_rates['costo_capital'],
            'tasa_carta_fianza': latest_rates['tasa_carta_fianza'],
            'captured_at': datetime.now(timezone.utc).isoformat(),
        }
        return latest_rates, snapshot

    def _extract_header_data(
        self,
        worksheet: Worksheet,
        latest_rates: _MasterRates,
        snapshot: _MasterSnapshot,
    ) -> _HeaderDict:
        """Read header cells, normalize keys, and inject master variables.

        Args:
            worksheet: The active openpyxl worksheet.
            latest_rates: Validated master variable rates (snake_case keys).
            snapshot: Frozen master variables snapshot for audit trail.

        Returns:
            Enriched header dict with snake_case keys and injected rates.
        """
        header_data: dict[str, Union[int, float, str]] = {}
        for var_name, cell_ref in self._config.VARIABLES_TO_EXTRACT.items():
            cell_value = worksheet[cell_ref].value

            if var_name in _DECIMAL_FIELDS:
                header_data[var_name] = safe_decimal(cell_value, self._logger)
            elif var_name in _INT_FIELDS:
                header_data[var_name] = safe_int(cell_value)
            elif var_name in _BOOL_FIELDS:
                header_data[var_name] = bool(cell_value) if cell_value is not None else False
            else:
                header_data[var_name] = str(cell_value) if cell_value is not None else ""

        header_data = normalize_keys(header_data)

        # The real commission is calculated by the financial engine later.
        if 'comisiones' in header_data:
            header_data['comisiones'] = Decimal("0")

        # Inject master variables into header data
        header_data['tipo_cambio'] = latest_rates['tipo_cambio']
        header_data['costo_capital_anual'] = latest_rates['costo_capital']
        header_data['tasa_carta_fianza'] = latest_rates['tasa_carta_fianza']
        header_data['aplica_carta_fianza'] = False
        header_data['master_variables_snapshot'] = snapshot

        return header_data

    def _extract_table_rows(
        self,
        worksheet: Worksheet,
        start_row: int,
        columns: dict[str, str],
    ) -> list[_RowDict]:
        """Extract rows from a worksheet table region with empty-row sentinel.

        Generic extraction used for both recurring services and fixed costs.
        Iterates from ``start_row + 1`` (skipping the header row) until
        either the worksheet ends or 5 consecutive empty rows are encountered.

        Args:
            worksheet: The active openpyxl worksheet.
            start_row: Config-defined start row (0-based header offset).
            columns: Mapping of ``{field_name: column_letter}``.

        Returns:
            List of normalized (snake_case) row dictionaries.
        """
        max_empty_rows: int = self._config.MAX_EMPTY_ROWS
        rows: list[_RowDict] = []
        empty_row_count: int = 0

        for row_idx in range(start_row + 1, worksheet.max_row + 1):
            row_data: _RowDict = {}
            is_empty_row: bool = True

            for field_name, col_letter in columns.items():
                col_idx: int = column_index_from_string(col_letter)
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value

                if cell_value is not None and str(cell_value).strip() != '':
                    is_empty_row = False

                row_data[field_name] = cell_value

            if is_empty_row:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    break
            else:
                empty_row_count = 0
                row_data = normalize_keys(row_data)
                rows.append(row_data)

        return rows

    def _transform_and_enrich(
        self,
        header_data: _HeaderDict,
        recurring_services_data: list[_RowDict],
        fixed_costs_data: list[_RowDict],
    ) -> Union[ServiceResult, dict[str, object]]:
        """Apply currency tagging, preview calculations, and validation.

        Enriches fixed costs with ``_original`` / ``total`` fields, recurring
        services with ``quantity`` / ``price_original`` / ``ingreso`` / ``egreso``
        fields, validates required header fields, and consolidates everything
        into a single data dict.

        Args:
            header_data: Enriched header dict from ``_extract_header_data``.
            recurring_services_data: Raw recurring service rows.
            fixed_costs_data: Raw fixed cost rows.

        Returns:
            On success: the consolidated ``full_extracted_data`` dict.
            On failure: a ``ServiceResult`` with a validation error.
        """
        # --- Fixed costs: currency tagging & preview totals ---
        for item in fixed_costs_data:
            item['costo_unitario_original'] = safe_decimal(item.get('costo_unitario', 0), self._logger)
            item['costo_unitario_currency'] = 'USD'

            cantidad = item.get('cantidad')
            costo_original = item.get('costo_unitario_original')
            if cantidad is not None and costo_original is not None:
                item['total'] = cantidad * costo_original

            item['periodo_inicio'] = safe_decimal(item.get('periodo_inicio', 0), self._logger)
            item['duracion_meses'] = safe_decimal(item.get('duracion_meses', 1), self._logger)

        # --- Recurring services: currency tagging & preview totals ---
        for item in recurring_services_data:
            q: Decimal = safe_decimal(item.get('q', 0), self._logger)
            p_original: Decimal = safe_decimal(item.get('p', 0), self._logger)
            cu1_original: Decimal = safe_decimal(item.get('cu1', 0), self._logger)
            cu2_original: Decimal = safe_decimal(item.get('cu2', 0), self._logger)

            item['quantity'] = q
            item['price_original'] = p_original
            item['price_currency'] = 'PEN'
            item['cost_unit_1_original'] = cu1_original
            item['cost_unit_2_original'] = cu2_original
            item['cost_unit_currency'] = 'USD'

            item['ingreso'] = q * p_original
            item['egreso'] = (cu1_original + cu2_original) * q

        # Total installation cost in original currency
        calculated_costo_instalacion: Decimal = sum(
            (item.get('total', Decimal("0")) for item in fixed_costs_data if item.get('total') is not None),
            Decimal("0"),
        )

        # Validate required fields
        client_name: Union[str, None] = header_data.get('client_name')
        mrc_value: Union[float, None] = header_data.get('mrc')
        if not client_name or client_name == '' or mrc_value is None or mrc_value == '':
            return ServiceResult(
                success=False,
                error="Required field 'Client Name' or 'MRC' is missing from the Excel file.",
                status_code=400,
            )

        # Rename to _original pattern for transaction
        header_data['mrc_original'] = header_data.get('mrc')
        header_data['mrc_currency'] = 'PEN'
        header_data['nrc_original'] = header_data.get('nrc')
        header_data['nrc_currency'] = 'PEN'

        # Consolidate all extracted data
        full_extracted_data: dict[str, object] = {
            **header_data,
            'recurring_services': recurring_services_data,
            'fixed_costs': fixed_costs_data,
            'costo_instalacion': calculated_costo_instalacion,
        }
        return full_extracted_data

    def _assemble_response(
        self,
        header_data: _HeaderDict,
        full_extracted_data: dict[str, object],
        fixed_costs_data: list[_RowDict],
        recurring_services_data: list[_RowDict],
        file_hash: str,
    ) -> ServiceResult:
        """Run financial engine and assemble the final JSON-safe response.

        Args:
            header_data: Enriched header dict.
            full_extracted_data: Consolidated data dict for the financial engine.
            fixed_costs_data: Enriched fixed cost rows.
            recurring_services_data: Enriched recurring service rows.
            file_hash: SHA-256 digest for chain-of-custody tracking.

        Returns:
            ``ServiceResult`` with the clean, JSON-safe data package.
        """
        financial_metrics: dict[str, object] = calculate_financial_metrics(
            full_extracted_data,
        ).model_dump()

        transaction_summary: dict[str, object] = {
            **header_data,
            **financial_metrics,
            "costo_instalacion": financial_metrics.get('costo_instalacion'),
            "submission_date": None,
            "approval_status": "PENDING",
            "file_sha256": file_hash,
        }

        final_data_package: dict[str, object] = {
            "transactions": transaction_summary,
            "fixed_costs": fixed_costs_data,
            "recurring_services": recurring_services_data,
        }

        clean_data: object = convert_to_json_safe(final_data_package)

        if isinstance(clean_data, dict):
            clean_data["file_sha256"] = file_hash

        return ServiceResult(success=True, data=clean_data)
